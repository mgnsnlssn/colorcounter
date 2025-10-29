from openpyxl import load_workbook, Workbook
from pathlib import Path
import datetime, os

INBOX = "inbox"
OUTBOX = "outbox"

# Färgkoder
RED_CODES = ["FF0000", "FF6666", "FFC7CE"]
YELLOW_CODES = ["FFFF00", "FFF200", "FFEB9C"]
GREEN_CODES = ["00FF00", "92D050"]

DAY_KEYWORDS = {
    "mån": "Monday",
    "tis": "Tuesday",
    "ons": "Wednesday",
    "tor": "Thursday",
    "fre": "Friday"
}

def get_rgb(cell):
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb:
        return fill.start_color.rgb[-6:].upper()
    return None

def detect_transitions(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    # --- Hitta dagkolumner (blockvis) ---
    day_columns = {}
    last_day = None
    for idx, cell in enumerate(ws[1], start=1):
        val = str(cell.value).lower() if cell.value else ""
        matched = False
        for sw, en in DAY_KEYWORDS.items():
            if sw in val:
                last_day = en
                day_columns.setdefault(en, []).append(idx)
                matched = True
                break
        if not matched and last_day:
            day_columns[last_day].append(idx)

    results = []

    for row in ws.iter_rows(min_row=2):
        name = row[0].value
        if not name:
            continue

        for day, cols in day_columns.items():
            day_cells = [row[i - 1] for i in cols]
            colors = [get_rgb(c) for c in day_cells]

            for i in range(len(colors) - 1):
                if (colors[i] in YELLOW_CODES and colors[i + 1] in RED_CODES):
                    results.append((name, day, "Yellow → Red"))
                    break
                elif (colors[i] in GREEN_CODES and colors[i + 1] in RED_CODES):
                    results.append((name, day, "Green → Red"))
                    break

    return results


def main():
    os.makedirs(OUTBOX, exist_ok=True)
    files = [f for f in os.listdir(INBOX) if f.endswith(".xlsx")]
    if not files:
        return

    for f in files:
        path = os.path.join(INBOX, f)
        results = detect_transitions(path)
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        out_path = os.path.join(OUTBOX, f"skolk_report_{Path(f).stem}_{timestamp}.xlsx")

        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Detected transitions"
        ws_out.append(["Student", "Day", "Transition"])

        if not results:
            ws_out.append(["No transitions detected", "", ""])
        else:
            for r in results:
                ws_out.append(r)

        wb_out.save(out_path)
        print(f"✅ Report saved → {out_path}")


if __name__ == "__main__":
    main()