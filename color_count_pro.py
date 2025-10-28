#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =====================================================
#  Color Count Pro v3.0
#  By: Mange & ChatGPT
#  - Räknar färger (grön/gul/röd) i Excel-filer
#  - Skapar statistik per text, vecka och totaltrend
# =====================================================

from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.colors import ColorChoice
from pathlib import Path
import re, time

# === Konfiguration ===
START_COL, END_COL = 1, 30
OUT_GREEN_COL, OUT_YELLOW_COL, OUT_RED_COL = 31, 32, 33
HEADER_ROW = 1

# === Färgkalibrering (från din fil) ===
TARGET_GREENS  = [(0, 204, 0)]      # Sen ankomst
TARGET_YELLOWS = [(255, 255, 51)]   # Sjuk
TARGET_REDS    = [(255, 51, 51)]    # Skolk
TOLERANCE = 40

INBOX  = Path("inbox")
OUTBOX = Path("outbox")
SUMMARY_FILE = Path("Sammanställning.xlsx")
SLEEP_SECONDS = 3

# === Hjälpfunktioner ===
def color_distance(c1,c2): return sum((a-b)**2 for a,b in zip(c1,c2))**0.5
def closest(rgb,palette): return min(color_distance(rgb,p) for p in palette)

def classify_rgb(rgb):
    if not rgb: return None
    d_r = closest(rgb,TARGET_REDS)
    d_g = closest(rgb,TARGET_GREENS)
    d_y = closest(rgb,TARGET_YELLOWS)
    best = min(d_r,d_g,d_y)
    if best>TOLERANCE: return None
    if best==d_g: return "green"
    if best==d_y: return "yellow"
    if best==d_r: return "red"

def get_rgb(cell):
    fill=cell.fill
    if not fill: return None
    color=getattr(fill,"fgColor",None) or getattr(fill,"start_color",None)
    if not color: return None
    rgb=getattr(color,"rgb",None)
    if rgb:
        try: return tuple(int(rgb[-6:][i:i+2],16) for i in (0,2,4))
        except: return None
    return None

def safe_set(ws,row,col,val):
    try: ws.cell(row=row,column=col,value=val)
    except AttributeError: ws.cell(row=row+1,column=col,value=val)

# === Bearbeta en fil ===
def process_file(path,out_dir):
    wb=load_workbook(path)
    ws=wb.active

    safe_set(ws,HEADER_ROW,OUT_GREEN_COL,"Antal_Gröna")
    safe_set(ws,HEADER_ROW,OUT_YELLOW_COL,"Antal_Gula")
    safe_set(ws,HEADER_ROW,OUT_RED_COL,"Antal_Röda")

    for r in range(HEADER_ROW+1,ws.max_row+1):
        g=y=rö=0
        for c in range(START_COL,END_COL+1):
            lbl=classify_rgb(get_rgb(ws.cell(row=r,column=c)))
            if lbl=="green": g+=1
            elif lbl=="yellow": y+=1
            elif lbl=="red": rö+=1
        if g+y+rö>0:
            ws.cell(row=r,column=OUT_GREEN_COL,value=g)
            ws.cell(row=r,column=OUT_YELLOW_COL,value=y)
            ws.cell(row=r,column=OUT_RED_COL,value=rö)

    # === Statistik per text ===
    stats={"green":{},"yellow":{},"red":{}}
    for r in range(HEADER_ROW+1,ws.max_row+1):
        for c in range(START_COL,END_COL+1):
            cell=ws.cell(row=r,column=c)
            text=str(cell.value).strip() if cell.value else None
            lbl=classify_rgb(get_rgb(cell))
            if lbl and text:
                stats[lbl][text]=stats[lbl].get(text,0)+1

    # === Statistikblad ===
    sheet_name=f"Statistik_{Path(path).stem.split('_')[0]}"
    if sheet_name in wb.sheetnames: del wb[sheet_name]
    ws_stat=wb.create_sheet(title=sheet_name)
    ws_stat.append(["Färg","Text","Antal"])
    for color_key, color_dict in stats.items():
        for text,count in sorted(color_dict.items(),key=lambda x:x[1],reverse=True):
            färg="Grön" if color_key=="green" else "Gul" if color_key=="yellow" else "Röd"
            ws_stat.append([färg,text,count])

    # Diagram för textstatistik
    if ws_stat.max_row>2:
        data_ref=Reference(ws_stat,min_col=3,min_row=1,max_row=ws_stat.max_row)
        cats_ref=Reference(ws_stat,min_col=2,min_row=2,max_row=ws_stat.max_row)
        chart=BarChart()
        chart.title="Färgstatistik per text"
        chart.add_data(data_ref,titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height=10; chart.width=22
        ws_stat.add_chart(chart,"E2")

    out_dir.mkdir(exist_ok=True)
    outpath=out_dir/(path.stem+"_with_counts.xlsx")
    wb.save(outpath)
    print(f"✅ {outpath.name} + Statistikflik klar")
    update_summary(outpath)
    return outpath

# === Uppdatera Sammanställning ===
def update_summary(processed_file):
    pat=re.compile(r"([a-zA-Z0-9\-]+)[_\-]v?w?(\d+)",re.I)
    m=pat.search(processed_file.stem)
    if not m:
        print(f"⚠️  {processed_file.name} matchar ej klass/vecka"); return
    klass,vecka=m.groups()

    wb_in=load_workbook(processed_file)
    ws_in=wb_in.active

    if SUMMARY_FILE.exists():
        wb_sum=load_workbook(SUMMARY_FILE)
    else:
        wb_sum=Workbook(); wb_sum.remove(wb_sum.active)

    blad=f"v{vecka}"
    if blad in wb_sum.sheetnames:
        ws_sum=wb_sum[blad]
    else:
        ws_sum=wb_sum.create_sheet(title=blad)
        ws_sum.append(["Klass","Elev","Gröna","Gula","Röda"])

    for row in ws_in.iter_rows(min_row=2,max_col=OUT_RED_COL):
        elev=str(row[0].value).strip() if row[0].value else ""
        g=row[OUT_GREEN_COL-1].value
        y=row[OUT_YELLOW_COL-1].value
        r=row[OUT_RED_COL-1].value
        if elev: ws_sum.append([klass,elev,g,y,r])

    # --- Extra: samla textstatistik globalt ---
    if "Statistik_Total" not in wb_sum.sheetnames:
        ws_total=wb_sum.create_sheet("Statistik_Total")
        ws_total.append(["Vecka","Färg","Text","Antal"])
    else:
        ws_total=wb_sum["Statistik_Total"]

    stat_sheet=[n for n in wb_in.sheetnames if n.startswith("Statistik_")]
    if stat_sheet:
        ws_stat_in=wb_in[stat_sheet[0]]
        for row in ws_stat_in.iter_rows(min_row=2,values_only=True):
            färg,text,antal=row
            if text and antal:
                ws_total.append([vecka,färg,text,antal])

    # --- Trender över veckor ---
    totals={}
    for row in ws_total.iter_rows(min_row=2,values_only=True):
        vecka,färg,text,antal=row
        if not isinstance(antal,(int,float)): continue
        totals.setdefault(vecka,{"Grön":0,"Gul":0,"Röd":0})
        totals[vecka][färg]=totals[vecka].get(färg,0)+antal

    for old in list(ws_total._charts): ws_total._charts.remove(old)
    start_row=ws_total.max_row+3
    ws_total.cell(row=start_row,column=1,value="Vecka")
    ws_total.cell(row=start_row,column=2,value="Grön")
    ws_total.cell(row=start_row,column=3,value="Gul")
    ws_total.cell(row=start_row,column=4,value="Röd")

    for i,(v,d) in enumerate(sorted(totals.items(),key=lambda x:int(x[0]))):
        ws_total.cell(row=start_row+1+i,column=1,value=int(v))
        ws_total.cell(row=start_row+1+i,column=2,value=d["Grön"])
        ws_total.cell(row=start_row+1+i,column=3,value=d["Gul"])
        ws_total.cell(row=start_row+1+i,column=4,value=d["Röd"])

    data_ref=Reference(ws_total,min_col=2,max_col=4,min_row=start_row,max_row=start_row+len(totals))
    cats_ref=Reference(ws_total,min_col=1,min_row=start_row+1,max_row=start_row+len(totals))
    chart=LineChart()
    chart.title="Trender över veckor"
    chart.y_axis.title="Antal"
    chart.x_axis.title="Vecka"
    chart.add_data(data_ref,titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height=10; chart.width=22

    # --- Färgmatchning för linjer ---
    for i,series in enumerate(chart.series):
        if i==0: series.graphicalProperties.line.solidFill="00CC00"   # Grön
        elif i==1: series.graphicalProperties.line.solidFill="FFFF33" # Gul
        elif i==2: series.graphicalProperties.line.solidFill="FF3333" # Röd

    ws_total.add_chart(chart,f"G{start_row}")

    wb_sum.save(SUMMARY_FILE)
    print(f"📊  Sammanställning uppdaterad + diagram för v{vecka}")

# === Bevakningsloop ===
def watch_loop():
    seen={}
    INBOX.mkdir(exist_ok=True)
    OUTBOX.mkdir(exist_ok=True)
    print(f"👀  Bevakar {INBOX.resolve()}")
    while True:
        for f in INBOX.glob("*.xlsx"):
            try: m=f.stat().st_mtime
            except FileNotFoundError: continue
            if f not in seen or seen[f]!=m:
                seen[f]=m
                process_file(f,OUTBOX)
        time.sleep(SLEEP_SECONDS)

if __name__=="__main__":
    watch_loop()