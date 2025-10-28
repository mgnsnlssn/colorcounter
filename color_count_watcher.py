#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import time
import argparse
from pathlib import Path
from openpyxl import load_workbook

# ========= KONFIG =========
# Kolumnintervall att räkna (A..Z = 1..26)
START_COL = 1      # A
END_COL   = 26     # Z

# Skrivresultat
OUT_GREEN_COL = 27  # AA
OUT_RED_COL   = 28  # AB
HEADER_ROW    = 1   # Antas vara rubrikrad

# Färgmatchning: målnyanser (RGB) och tolerans (0..441 ungefär)
# Här ligger nyanser från din fil: grön (0,204,0), röd (255,51,51)
TARGET_GREENS = [(0, 204, 0), (0, 255, 0), (146, 208, 80)]
TARGET_REDS   = [(255, 51, 51), (255, 0, 0), (230, 92, 92)]
TOLERANCE     = 60  # större = mer generös matchning

# Mappbevakning
INBOX  = Path("inbox")
OUTBOX = Path("outbox")
SLEEP_SECONDS = 2
# ==========================


def color_distance(c1, c2):
    # euklidiskt avstånd i RGB
    return ((c1[0]-c2[0])**2 + (c1[1]-c2[1])**2 + (c1[2]-c2[2])**2) ** 0.5


def closest_match(rgb, palette):
    return min((color_distance(rgb, p), p) for p in palette)[0]


def get_rgb_from_fill(cell):
    """
    Försök läsa bakgrundsfärg som RGB-tuple (r,g,b).
    Hanterar vanliga ARGB, fallback för indexed.
    """
    fill = cell.fill
    if not fill:
        return None

    # Försök fgColor först
    color = getattr(fill, "fgColor", None)
    if not color or (getattr(color, "rgb", None) is None and getattr(color, "indexed", None) is None):
        # Ibland används start_color
        color = getattr(fill, "start_color", None)

    if not color:
        return None

    # Direkt RGB (ARGB 'FFRRGGBB' eller 'RRGGBB')
    rgb = getattr(color, "rgb", None)
    if rgb:
        hex6 = rgb[-6:]
        try:
            r = int(hex6[0:2], 16)
            g = int(hex6[2:4], 16)
            b = int(hex6[4:6], 16)
            return (r, g, b)
        except Exception:
            pass

    # Indexed palette (mycket basic fallback)
    idx = getattr(color, "indexed", None)
    if idx is not None:
        palette = {
            0: (0, 0, 0),
            1: (255, 255, 255),
            2: (255, 0, 0),
            3: (0, 255, 0),
            4: (0, 0, 255),
            5: (255, 255, 0),
            6: (255, 0, 255),
            7: (0, 255, 255),
            10: (128, 0, 0),
            11: (0, 128, 0),
            12: (0, 0, 128),
            15: (192, 192, 192),
            16: (128, 128, 128),
            17: (255, 0, 0),
            18: (0, 255, 0),
            19: (0, 0, 255),
        }
        return palette.get(idx, None)

    # Themefärger hanteras ej här
    return None


def classify_rgb(rgb):
    if not rgb:
        return None
    # matcha mot närmaste mål i respektive palett
    d_red   = closest_match(rgb,   TARGET_REDS)
    d_green = closest_match(rgb, TARGET_GREENS)
    if d_red <= TOLERANCE and d_red < d_green:
        return "red"
    if d_green <= TOLERANCE and d_green < d_red:
        return "green"
    return None


def safe_set(ws, row, col, value):
    """Skriv värde, hoppa ner en rad om cellen är en del av en merge (read-only)."""
    try:
        ws.cell(row=row, column=col, value=value)
    except AttributeError:
        ws.cell(row=row+1, column=col, value=value)


def process_file(input_path: Path, out_dir: Path) -> Path:
    wb = load_workbook(str(input_path))
    ws = wb.active

    # Rubriker
    safe_set(ws, HEADER_ROW, OUT_GREEN_COL, "Antal_Gröna")
    safe_set(ws, HEADER_ROW, OUT_RED_COL,   "Antal_Röda")

    max_row = ws.max_row

    for r in range(HEADER_ROW + 1, max_row + 1):
        green_count = 0
        red_count = 0
        row_has_data = False

        for c in range(START_COL, END_COL + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value not in (None, ""):
                row_has_data = True
            rgb = get_rgb_from_fill(cell)
            label = classify_rgb(rgb)
            if label == "green":
                green_count += 1
            elif label == "red":
                red_count += 1

        if row_has_data or (green_count + red_count) > 0:
            ws.cell(row=r, column=OUT_GREEN_COL, value=green_count)
            ws.cell(row=r, column=OUT_RED_COL,   value=red_count)

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / (input_path.stem + "_with_counts.xlsx")
    wb.save(str(out_path))
    return out_path


def watch_loop(inbox: Path, outbox: Path):
    print(f"Watching: {inbox.resolve()} → {outbox.resolve()}")
    inbox.mkdir(parents=True, exist_ok=True)
    outbox.mkdir(parents=True, exist_ok=True)

    seen = {}  # path -> last_mtime

    while True:
        for p in inbox.glob("*.xlsx"):
            try:
                mtime = p.stat().st_mtime
            except FileNotFoundError:
                continue

            # Om ny eller ändrad → processa
            if p not in seen or seen[p] != mtime:
                print(f"Processing: {p.name}")
                try:
                    out_path = process_file(p, outbox)
                    print(f"  → OK: {out_path.name}")
                    seen[p] = mtime
                except Exception as e:
                    print(f"  ! ERROR on {p.name}: {e}")
        time.sleep(SLEEP_SECONDS)


def main():
    ap = argparse.ArgumentParser(description="Räkna röda/gröna Excel-celler och skriv resultat i AA/AB.")
    ap.add_argument("--once", help="Kör på en enda fil och avsluta (path till .xlsx).")
    ap.add_argument("--watch", action="store_true", help="Bevaka 'inbox/' och skriv resultat till 'outbox/'.")
    args = ap.parse_args()

    if args.once:
        in_file = Path(args.once)
        if not in_file.exists():
            raise SystemExit(f"Hittar inte filen: {in_file}")
        out = process_file(in_file, Path("."))
        print(f"KLART → {out}")
        return

    if args.watch:
        watch_loop(INBOX, OUTBOX)
        return

    ap.print_help()


if __name__ == "__main__":
    main()